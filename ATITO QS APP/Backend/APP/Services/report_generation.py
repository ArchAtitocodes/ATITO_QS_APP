# backend/app/services/report_generation.py
"""
Report Generation Service
Generates BoQ, BBS, and other reports in Excel and PDF formats
Author: Eng. STEPHEN ODHIAMBO
"""

from typing import Dict, List, Any, BinaryIO
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import io
from datetime import datetime

from app.models.project import Project
from app.models.boq import BOQItem
from app.models.bbs import BBSItem
from app.config import settings


class ReportService:
    """
    Service for generating professional reports
    """
    
    def __init__(self, project: Project, db: Session):
        self.project = project
        self.db = db
    
    # ========== EXCEL REPORTS ==========
    
    def generate_boq_excel(self) -> io.BytesIO:
        """
        Generate Bill of Quantities in Excel format
        Professional formatting with formulas
        """
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bill of Quantities"
        
        # Set column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        # Title section
        title_font = Font(name='Arial', size=16, bold=True)
        header_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)
        
        # Header fill
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = "BILL OF QUANTITIES"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:G1')
        
        # Project details
        row = 3
        ws[f'A{row}'] = "Project:"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = self.project.name
        
        row += 1
        ws[f'A{row}'] = "Location:"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = f"{self.project.location}, {self.project.county}"
        
        row += 1
        ws[f'A{row}'] = "Date:"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = datetime.utcnow().strftime("%d %B %Y")
        
        row += 1
        ws[f'A{row}'] = "Client Type:"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = self.project.client_type or "N/A"
        
        row += 2
        
        # Column headers
        headers = ['Item No.', 'Description', 'Unit', 'Quantity', 'Rate (KES)', 'Amount (KES)', 'Remarks']
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row += 1
        start_data_row = row
        
        # Fetch BoQ items
        boq_items = self.db.query(BOQItem).filter(
            BOQItem.project_id == self.project.id
        ).order_by(BOQItem.item_number).all()
        
        # Group by category
        current_category = None
        category_start_rows = {}
        
        for item in boq_items:
            # Category header
            if item.category != current_category:
                current_category = item.category
                category_start_rows[current_category] = row
                
                # Category row
                ws.merge_cells(f'A{row}:G{row}')
                cat_cell = ws[f'A{row}']
                cat_cell.value = f"═══ {current_category.upper()} ═══"
                cat_cell.font = Font(bold=True, size=11)
                cat_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cat_cell.alignment = Alignment(horizontal='center')
                row += 1
            
            # Item data
            ws[f'A{row}'] = item.item_number
            ws[f'B{row}'] = item.description
            ws[f'C{row}'] = item.unit
            ws[f'D{row}'] = round(item.gross_quantity, 2)
            ws[f'E{row}'] = round(item.unit_rate, 2)
            ws[f'F{row}'] = f"=D{row}*E{row}"  # Formula for amount
            ws[f'G{row}'] = item.remarks or ""
            
            # Apply borders
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border
            
            # Number formatting
            ws[f'D{row}'].number_format = '#,##0.00'
            ws[f'E{row}'].number_format = '#,##0.00'
            ws[f'F{row}'].number_format = '#,##0.00'
            
            # Confidence indicator
            if item.needs_review:
                ws[f'G{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            row += 1
        
        # Subtotal row
        row += 1
        ws[f'E{row}'] = "SUBTOTAL:"
        ws[f'E{row}'].font = header_font
        ws[f'F{row}'] = f"=SUM(F{start_data_row}:F{row-2})"
        ws[f'F{row}'].font = header_font
        ws[f'F{row}'].number_format = '#,##0.00'
        
        # Cost summary section
        from app.services.costing_engine import CostingEngine
        costing = CostingEngine(self.project, self.db)
        cost_summary = costing.calculate_final_cost()
        
        row += 2
        summary_start = row
        
        summary_items = [
            ("Materials Subtotal", cost_summary['materials_subtotal']),
            ("Preliminary Cost (5%)", cost_summary['preliminary_cost']),
            ("Provisional Sum (10%)", cost_summary['provisional_sum']),
            ("Labor & Overheads (50%)", cost_summary['labor_overheads']),
            ("Subtotal", cost_summary['subtotal_before_contingency']),
            (f"Contingency ({cost_summary['contingency_percentage']:.0f}%)", cost_summary['contingency_amount']),
            ("Subtotal (Before VAT)", cost_summary['subtotal_before_vat']),
            (f"VAT ({cost_summary['vat_percentage']:.0f}%)", cost_summary['vat_amount']),
        ]
        
        for label, amount in summary_items:
            ws[f'E{row}'] = label
            ws[f'E{row}'].font = normal_font
            ws[f'F{row}'] = amount
            ws[f'F{row}'].number_format = '#,##0.00'
            row += 1
        
        # Grand total
        ws[f'E{row}'] = "GRAND TOTAL:"
        ws[f'E{row}'].font = Font(bold=True, size=12)
        ws[f'F{row}'] = cost_summary['grand_total']
        ws[f'F{row}'].font = Font(bold=True, size=12)
        ws[f'F{row}'].number_format = '#,##0.00'
        ws[f'F{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        # Footer
        row += 3
        ws[f'A{row}'] = "Prepared by: ATITO QS App"
        ws[f'A{row}'].font = Font(italic=True, size=9)
        row += 1
        ws[f'A{row}'] = f"Author: Eng. STEPHEN ODHIAMBO - Civil Engineer & AI Engineer"
        ws[f'A{row}'].font = Font(italic=True, size=9, bold=True)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def generate_bbs_excel(self) -> io.BytesIO:
        """
        Generate Bar Bending Schedule in Excel format
        BS 8666 compliant formatting
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bar Bending Schedule"
        
        # Set column widths
        column_widths = {
            'A': 10, 'B': 15, 'C': 20, 'D': 8, 'E': 8,
            'F': 10, 'G': 8, 'H': 8, 'I': 8, 'J': 8,
            'K': 8, 'L': 12, 'M': 10, 'N': 12, 'O': 20
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Styling
        title_font = Font(name='Arial', size=16, bold=True)
        header_font = Font(name='Arial', size=10, bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = "BAR BENDING SCHEDULE (BS 8666:2005)"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:O1')
        
        # Project details
        row = 3
        ws[f'A{row}'] = "Project:"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = self.project.name
        ws.merge_cells(f'B{row}:E{row}')
        
        row += 1
        ws[f'A{row}'] = "Date:"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = datetime.utcnow().strftime("%d %B %Y")
        
        row += 2
        
        # Column headers
        headers = [
            'Bar Mark', 'Member', 'Location', 'Dia.', 'Type',
            'Shape Code', 'A', 'B', 'C', 'D', 'E',
            'Total Length (mm)', 'No. of Bars', 'Total Weight (kg)', 'Remarks'
        ]
        
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
        
        row += 1
        start_data_row = row
        
        # Fetch BBS items
        bbs_items = self.db.query(BBSItem).filter(
            BBSItem.project_id == self.project.id
        ).order_by(BBSItem.member_type, BBSItem.bar_mark).all()
        
        # Group by member type
        current_member = None
        
        for item in bbs_items:
            # Member type header
            if item.member_type != current_member:
                current_member = item.member_type
                
                ws.merge_cells(f'A{row}:O{row}')
                member_cell = ws[f'A{row}']
                member_cell.value = f"═══ {current_member.upper()} ═══"
                member_cell.font = Font(bold=True, size=11)
                member_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                member_cell.alignment = Alignment(horizontal='center')
                row += 1
            
            # Item data
            data_columns = [
                item.bar_mark,
                item.member_type,
                item.member_location or "",
                item.bar_diameter,
                item.bar_type,
                item.shape_code,
                item.length_a or "",
                item.length_b or "",
                item.length_c or "",
                item.length_d or "",
                item.length_e or "",
                item.total_length,
                item.number_of_bars,
                item.total_weight,
                item.remarks or ""
            ]
            
            for col_num, value in enumerate(data_columns, start=1):
                cell = ws.cell(row=row, column=col_num)
                cell.value = value
                cell.border = thin_border
                
                # Number formatting
                if col_num >= 7 and col_num <= 14 and value:
                    cell.number_format = '#,##0.00'
            
            row += 1
        
        # Total steel weight
        row += 1
        ws[f'M{row}'] = "TOTAL STEEL:"
        ws[f'M{row}'].font = Font(bold=True)
        ws[f'N{row}'] = f"=SUM(N{start_data_row}:N{row-2})"
        ws[f'N{row}'].font = Font(bold=True)
        ws[f'N{row}'].number_format = '#,##0.00'
        ws[f'O{row}'] = "kg"
        
        # Footer
        row += 3
        ws[f'A{row}'] = "Standard: BS 8666:2005 - Scheduling, dimensioning, bending and cutting of steel reinforcement"
        ws[f'A{row}'].font = Font(italic=True, size=9)
        ws.merge_cells(f'A{row}:O{row}')
        
        row += 1
        ws[f'A{row}'] = "Prepared by: ATITO QS App | Author: Eng. STEPHEN ODHIAMBO"
        ws[f'A{row}'].font = Font(italic=True, size=9, bold=True)
        ws.merge_cells(f'A{row}:O{row}')
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    # ========== PDF REPORTS ==========
    
    def generate_boq_pdf(self) -> io.BytesIO:
        """
        Generate Bill of Quantities in PDF format
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12
        )
        
        # Title
        title = Paragraph("BILL OF QUANTITIES", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Project details
        project_info = [
            ['Project:', self.project.name],
            ['Location:', f"{self.project.location}, {self.project.county}"],
            ['Date:', datetime.utcnow().strftime("%d %B %Y")],
            ['Client Type:', self.project.client_type or "N/A"]
        ]
        
        info_table = Table(project_info, colWidths=[1.5*inch, 4.5*inch])
        info_table.setStyle(TableStyle([
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
            ('FONT', (1, 0), (1, -1), 'Helvetica', 10),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # BoQ table
        boq_items = self.db.query(BOQItem).filter(
            BOQItem.project_id == self.project.id
        ).order_by(BOQItem.item_number).all()
        
        # Table data
        table_data = [['Item No.', 'Description', 'Unit', 'Qty', 'Rate (KES)', 'Amount (KES)']]
        
        current_category = None
        for item in boq_items:
            # Category header
            if item.category != current_category:
                current_category = item.category
                table_data.append([
                    {'text': f"═══ {current_category.upper()} ═══", 'colspan': 6}
                ])
            
            # Item row
            table_data.append([
                item.item_number,
                item.description[:60] + '...' if len(item.description) > 60 else item.description,
                item.unit,
                f"{item.gross_quantity:,.2f}",
                f"{item.unit_rate:,.2f}",
                f"{item.total_cost:,.2f}"
            ])
        
        # Create table
        boq_table = Table(table_data, colWidths=[0.8*inch, 3*inch, 0.6*inch, 0.8*inch, 1*inch, 1.2*inch])
        boq_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            # Data
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        elements.append(boq_table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return buffer
    
    def generate_cost_summary_pdf(self) -> io.BytesIO:
        """
        Generate detailed cost summary PDF
        """
        from app.services.costing_engine import CostingEngine
        costing = CostingEngine(self.project, self.db)
        cost_summary = costing.generate_cost_summary()
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph("PROJECT COST SUMMARY", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Cost breakdown
        cost_data = [
            ['Description', 'Amount (KES)'],
            ['Materials Subtotal', f"{cost_summary['costs']['materials_subtotal']:,.2f}"],
            ['Preliminary Cost (5%)', f"{cost_summary['costs']['preliminary_cost']:,.2f}"],
            ['Provisional Sum (10%)', f"{cost_summary['costs']['provisional_sum']:,.2f}"],
            ['Labor & Overheads (50%)', f"{cost_summary['costs']['labor_overheads']:,.2f}"],
            ['Subtotal', f"{cost_summary['costs']['subtotal_before_contingency']:,.2f}"],
            ['Contingency', f"{cost_summary['costs']['contingency_amount']:,.2f}"],
            ['Subtotal (Before VAT)', f"{cost_summary['costs']['subtotal_before_vat']:,.2f}"],
            ['VAT (16%)', f"{cost_summary['costs']['vat_amount']:,.2f}"],
            ['GRAND TOTAL', f"{cost_summary['costs']['grand_total']:,.2f}"]
        ]
        
        cost_table = Table(cost_data, colWidths=[4*inch, 2*inch])
        cost_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.gold),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(cost_table)
        
        # Footer
        elements.append(Spacer(1, 0.5*inch))
        footer = Paragraph(
            "Prepared by: ATITO QS App | Author: <b>Eng. STEPHEN ODHIAMBO</b> - Civil Engineer & AI Engineer",
            styles['Normal']
        )
        elements.append(footer)
        
        doc.build(elements)
        buffer.seek(0)
        
        return buffer
